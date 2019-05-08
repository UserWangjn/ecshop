import time
# 导入操作excel文件的函数，如果没有安装，可通过pip install openpyxl安装
from openpyxl import workbook
from openpyxl import load_workbook
import os.path
from common.loggen import LogGen
from selenium import webdriver
from pages.mainpage import MainPage
from pages.basepage import BasePage
from pages.loginpage import LoginPage
from pages.registerpage import RegisterPage
from common.geturl import GetUrl

logger = LogGen(logger="TestSuite").getlog()


# 创建读取测试集函数
def read_testsuite(tsname):
    # 设置测试用例读取执行状态标志位
    flag = True
    # 设置读取测试集函数执行状态标志位
    read_testsuite = True
    # 判断测试集文件是否存在
    if os.path.exists(tsname):
        # 如果存在则写入日志
        logger.info('已找到TestSuite文件，开始分析测试集...')
        # 创建excel操作对象
        wbexcel = load_workbook(tsname)
        sheetnames = wbexcel.get_sheet_names()
        ws = wbexcel.get_sheet_by_name(sheetnames[0])
        # 分析测试集文件中的执行信息：执行标志位及测试脚本名称，从第二行开始
        for irow in range(2, ws.max_row + 1):
            # 获取测试集文件中的执行标志位值，从第二行开始，第二列
            testoperation = ws.cell(row=irow, column=2).value
            # 获取测试集文件中的测试用例名称，从第二行开始，第三列
            testcasefile = ws.cell(row=irow, column=3).value
            # 判断执行标志位是否需要执行，如果是do，则调用测试用例执行函数，如果是not,则
            # 不执行，如果是其xxxx值，则写入日志，报告执行参数错误，
            # 并指出是哪个用例执行参数错误
            if testoperation == 'do':
                logger.info('*********************')
                logger.info('执行 %s 测试场景' % testcasefile)
                # 加载测试用例读取函数，并返回其返回值，以判断用例读取情况
                flag = read_testcase(testcasefile)
                # 如果用例读取函数返回未False，则说明用例读取错误
                if flag == False:
                    logger.info('测试用例执行失败')
            # 如果执行状态为not，说明当前用例无须执行
            elif testoperation == 'not':
                logger.info('%s 场景无需测试' % testcasefile)
            # 如果既不是do，又不是not，则报告错误
            else:
                logger.info('执行参数错误，请检查%s' % testcasefile)
                # 如果执行状态错误，则跳出循环，停止测试
                break
    # 如果测试集文件错误，则写入日志，并提示错误原因
    else:
        logger.info('未发现:%s, 请检查文件是否正确' % tsname)
        # 返回测试集执行函数状态，便于run.py中的unittest中记录该状态
        read_testsuite = False
    # 返回测试集执行函数执行状态
    return read_testsuite


# 定义浏览器启动函数，本次并没有使用common中定义的browserlauncher函数，读者可自行扩展改写
def get_driver(testpage, teststep, testdata):
    # 设置浏览器启动函数执行状态，便于后续运行控制
    get_driver = True
    # 判断测试用例中是否需要启动浏览器，如果需要，则判断启动哪种浏览器
    if testpage == '浏览器':
        # 考虑测试用例中的step大小写问题，读者自行研究解决
        if teststep == 'firefox':
            driver = webdriver.Firefox()
        elif teststep == 'ie':
            driver = webdriver.Ie()
        elif teststep == 'chrome':
            driver = webdriver.Chrome()
        # 如果浏览器类型设置错误，写入日志并给予提示
        else:
            logger.info('未知浏览器类型，请检查测试用例')
        # 启动没有问题后加载测试路径并返回driver对象
        driver.get(testdata)
        get_driver = driver
    else:
        # 如果测试用例中的启动参数错误，则写入日志并给予提示
        logger.info('浏览器数据错误，请检查测试用例配置')
        get_driver = False
    return get_driver
